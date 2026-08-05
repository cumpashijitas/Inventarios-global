"""
Genera 5 scripts SQL INSERT desde INVENTARIO_UNIFICADO.xlsx.
Cada script inserta ~3,350 productos directamente en public.productos.
"""
import openpyxl
import math
import os

# ── Helpers de escape ─────────────────────────────────────────────────────────
def esc(v):
    if v is None or str(v).strip() == "":
        return "NULL"
    return "'" + str(v).strip().replace("'", "''") + "'"

def num(v, default="NULL"):
    if v is None or str(v).strip() == "":
        return default
    try:
        n = float(str(v))
        return str(round(n, 4)).rstrip("0").rstrip(".")
    except Exception:
        return default

# ── Leer Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook("backend/example_tables/INVENTARIO_UNIFICADO.xlsx", data_only=True)
ws = wb.active

rows = []
for r in range(2, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    if row[0]:  # SKU no vacío
        rows.append(row)

print(f"Total productos: {len(rows)}")

# ── Generar 5 archivos ────────────────────────────────────────────────────────
PARTS = 5
BATCH = math.ceil(len(rows) / PARTS)
os.makedirs("sql", exist_ok=True)

for part in range(PARTS):
    chunk = rows[part * BATCH:(part + 1) * BATCH]
    if not chunk:
        continue

    filename = f"sql/22-productos-parte-{part + 1}.sql"
    lines = []

    lines.append(f"-- =============================================================================")
    lines.append(f"-- 22-productos-parte-{part + 1}.sql  ({len(chunk)} productos)")
    lines.append(f"-- Parte {part + 1} de {PARTS} — ejecutar en Supabase SQL Editor")
    lines.append(f"-- =============================================================================")
    lines.append("")
    lines.append("DO $$")
    lines.append("DECLARE")
    lines.append("  v_eid  uuid;")
    lines.append("  v_uid  uuid;")
    lines.append("  v_cats jsonb;")
    lines.append("BEGIN")
    lines.append("  SELECT id INTO v_eid FROM public.empresas LIMIT 1;")
    lines.append("")
    lines.append("  SELECT id INTO v_uid")
    lines.append("    FROM public.unidades_medida")
    lines.append("   WHERE empresa_id = v_eid AND activo = true")
    lines.append("   ORDER BY created_at LIMIT 1;")
    lines.append("")
    lines.append("  -- Mapa categoria nombre -> uuid")
    lines.append("  SELECT jsonb_object_agg(upper(trim(nombre)), id::text)")
    lines.append("    INTO v_cats")
    lines.append("    FROM public.categorias")
    lines.append("   WHERE empresa_id = v_eid AND activo = true;")
    lines.append("")
    lines.append("  INSERT INTO public.productos (")
    lines.append("    empresa_id, sku, nombre, marca, codigo_universal, procedencia,")
    lines.append("    precio_compra, costo_caja, precio_venta, precio_mayor,")
    lines.append("    precio_mecanico, precio_real,")
    lines.append("    categoria_id, industria, motor, modelos, medidas,")
    lines.append("    stock_minimo, unidad_id, controla_stock, activo")
    lines.append("  ) VALUES")

    value_rows = []
    for row in chunk:
        sku         = row[0]
        cod_univ    = row[1]
        nombre      = row[2]
        marca       = row[3]
        procedencia = row[4]
        costo_unit  = row[5]
        costo_caja  = row[6]
        p_factura   = row[7]
        p_mayor     = row[8]
        p_taller    = row[9]
        p_mayorista = row[10]
        categoria   = row[11]
        industria   = row[12]
        motor       = row[13]
        modelos     = row[14]
        medidas     = row[15]
        # col 16 = proveedor (texto, sin FK directa en este insert)
        # col 17 = cantidad (siempre 0)
        stock_min   = row[18]
        # col 19 = unidad (texto, ya usamos v_uid)

        # Categoria: lookup en el JSON precargado
        if categoria and str(categoria).strip():
            cat_key = str(categoria).strip().upper().replace("'", "''")
            cat_expr = f"(v_cats->>'{cat_key}')::uuid"
        else:
            cat_expr = "NULL"

        val = (
            f"    (v_eid, {esc(sku)}, {esc(nombre)}, {esc(marca)}, "
            f"{esc(cod_univ)}, {esc(procedencia)},\n"
            f"     {num(costo_unit, '0')}, {num(costo_caja)}, "
            f"{num(p_factura, '0')}, {num(p_mayor)},\n"
            f"     {num(p_taller)}, {num(p_mayorista)},\n"
            f"     {cat_expr}, {esc(industria)}, {esc(motor)}, "
            f"{esc(modelos)}, {esc(medidas)},\n"
            f"     {num(stock_min, '0')}, v_uid, true, true)"
        )
        value_rows.append(val)

    lines.append(",\n".join(value_rows))
    lines.append("  ON CONFLICT (empresa_id, sku) DO UPDATE SET")
    lines.append("    precio_venta     = EXCLUDED.precio_venta,")
    lines.append("    precio_mayor     = EXCLUDED.precio_mayor,")
    lines.append("    precio_mecanico  = EXCLUDED.precio_mecanico,")
    lines.append("    precio_real      = EXCLUDED.precio_real,")
    lines.append("    precio_compra    = EXCLUDED.precio_compra,")
    lines.append("    costo_caja       = EXCLUDED.costo_caja;")
    lines.append("")
    lines.append(f"  RAISE NOTICE 'Parte {part + 1}/{PARTS} OK — {len(chunk)} productos procesados';")
    lines.append("END $$;")

    with open(filename, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    size_kb = os.path.getsize(filename) // 1024
    print(f"  Parte {part + 1}: {len(chunk)} productos -> {filename} ({size_kb} KB)")

print("\nListo.")
