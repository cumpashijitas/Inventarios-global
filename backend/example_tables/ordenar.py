import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("backend/example_tables/INVENTARIO_UNIFICADO.xlsx", data_only=True)
ws = wb.active

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
rows = []
for r in range(2, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    if any(v for v in row):
        rows.append(row)

# Ordenar por DESCRIPCION (columna 3, indice 2), A-Z, ignorar None
rows.sort(key=lambda r: str(r[2] if r[2] else "").strip().upper())

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Inventario Unificado"

hdr_fill = PatternFill("solid", fgColor="FFBB00")
hdr_font = Font(bold=True, color="000000", size=10)

for col, h in enumerate(headers, 1):
    c = ws_out.cell(1, col, h)
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(horizontal="center", wrap_text=True)

for row in rows:
    ws_out.append(row)

widths = [18, 18, 45, 12, 12, 10, 10, 10, 10, 10, 10, 22, 18, 12, 25, 12, 20, 8, 8, 6]
for i, w in enumerate(widths, 1):
    ws_out.column_dimensions[get_column_letter(i)].width = w

ws_out.freeze_panes = "A2"
wb_out.save("backend/example_tables/INVENTARIO_UNIFICADO.xlsx")

print(f"Ordenado. Total filas: {len(rows)}")
print(f"Primeras 3: {[r[2] for r in rows[:3]]}")
print(f"Ultimas 3:  {[r[2] for r in rows[-3:]]}")
