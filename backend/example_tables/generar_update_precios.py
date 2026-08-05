"""
Genera un único SQL que actualiza solo los precios de productos existentes.
Usa el SKU como clave. No inserta nuevos productos.
"""
import openpyxl
import math
import os

def num(v, default="NULL"):
    if v is None or str(v).strip() == "":
        return default
    try:
        n = float(str(v))
        return str(round(n, 4)).rstrip("0").rstrip(".")
    except Exception:
        return default

def esc(v):
    if v is None or str(v).strip() == "":
        return "NULL"
    return "'" + str(v).strip().replace("'", "''") + "'"

wb = openpyxl.load_workbook("backend/example_tables/INVENTARIO_UNIFICADO.xlsx", data_only=True)
ws = wb.active

rows = []
for r in range(2, ws.max_row + 1):
    sku         = ws.cell(r, 1).value
    p_factura   = ws.cell(r, 8).value
    p_mayor     = ws.cell(r, 9).value
    p_taller    = ws.cell(r, 10).value
    p_mayorista = ws.cell(r, 11).value
    costo_unit  = ws.cell(r, 6).value
    costo_caja  = ws.cell(r, 7).value
    if sku:
        rows.append((sku, p_factura, p_mayor, p_taller, p_mayorista, costo_unit, costo_caja))

print(f"Total productos a actualizar: {len(rows)}")

PARTS = 5
BATCH = math.ceil(len(rows) / PARTS)
os.makedirs("sql", exist_ok=True)

for part in range(PARTS):
    chunk = rows[part * BATCH:(part + 1) * BATCH]
    if not chunk:
        continue

    filename = f"sql/25-update-precios-parte-{part + 1}.sql"
    lines = []
    lines.append("-- =============================================================================")
    lines.append(f"-- 25-update-precios-parte-{part + 1}.sql  ({len(chunk)} productos)")
    lines.append(f"-- Parte {part + 1} de {PARTS} — actualiza SOLO precios, no inserta")
    lines.append("-- Ejecutar en Supabase SQL Editor")
    lines.append("-- =============================================================================")
    lines.append("")
    lines.append("DO $$")
    lines.append("DECLARE")
    lines.append("  v_eid uuid;")
    lines.append("  v_count int := 0;")
    lines.append("BEGIN")
    lines.append("  SELECT id INTO v_eid FROM public.empresas LIMIT 1;")
    lines.append("")

    for sku, p_factura, p_mayor, p_taller, p_mayorista, costo_unit, costo_caja in chunk:
        lines.append(
            f"  UPDATE public.productos SET"
            f" precio_venta = {num(p_factura, '0')},"
            f" precio_mayor = {num(p_mayor)},"
            f" precio_mecanico = {num(p_taller)},"
            f" precio_real = {num(p_mayorista)},"
            f" precio_compra = {num(costo_unit, '0')},"
            f" costo_caja = {num(costo_caja)}"
            f" WHERE empresa_id = v_eid AND sku = {esc(sku)};"
        )
        lines.append("  IF FOUND THEN v_count := v_count + 1; END IF;")

    lines.append("")
    lines.append(f"  RAISE NOTICE 'Parte {part + 1}/{PARTS} — Precios actualizados: % productos', v_count;")
    lines.append("END $$;")

    with open(filename, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    size_kb = os.path.getsize(filename) // 1024
    print(f"  Parte {part + 1}: {len(chunk)} productos -> {filename} ({size_kb} KB)")
