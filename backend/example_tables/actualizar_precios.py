import openpyxl

# Cargar precios de VENDEDOR indexados por CODIGO MARCA
wb_v = openpyxl.load_workbook("backend/example_tables/TABLA_VENDEDOR.xlsx", data_only=True)
ws_v = wb_v.active

# VENDEDOR cols: CODIGO MARCA=2, PRECIO FACTURA=9, POR MAYOR=10, TALLER=11, MAYORISTA=12
precios_vendedor = {}
for row in ws_v.iter_rows(min_row=2, values_only=True):
    codigo = row[1]  # CODIGO MARCA
    if codigo:
        precios_vendedor[str(codigo).strip().upper()] = {
            "factura":   row[8],
            "por_mayor": row[9],
            "taller":    row[10],
            "mayorista": row[11],
        }

print(f"Productos en VENDEDOR: {len(precios_vendedor)}")

# Actualizar UNIFICADO
wb_u = openpyxl.load_workbook("backend/example_tables/INVENTARIO_UNIFICADO.xlsx", data_only=True)
ws_u = wb_u.active

# UNIFICADO cols: CODIGO MARCA=1, PRECIO FACTURA=8, POR MAYOR=9, TALLER=10, MAYORISTA=11
actualizados = 0
sin_coincidencia = 0

for row in ws_u.iter_rows(min_row=2):
    codigo = row[0].value  # CODIGO MARCA (col 1)
    if not codigo:
        sin_coincidencia += 1
        continue

    key = str(codigo).strip().upper()
    if key in precios_vendedor:
        p = precios_vendedor[key]
        row[7].value  = p["factura"]    # col 8
        row[8].value  = p["por_mayor"]  # col 9
        row[9].value  = p["taller"]     # col 10
        row[10].value = p["mayorista"]  # col 11
        actualizados += 1
    else:
        sin_coincidencia += 1

wb_u.save("backend/example_tables/INVENTARIO_UNIFICADO.xlsx")

print(f"Precios actualizados desde VENDEDOR: {actualizados}")
print(f"Sin coincidencia (precios originales): {sin_coincidencia}")
